from selenium import webdriver
import openpyxl
driver = webdriver.Chrome(executable_path=r'C:\Users\shyam raj\untitled\atcost\drivers\chromedriver.exe')
path=r"https://d.docs.live.net/26c21ed15cc4a0b8/atcost.xlsx"
workbook=openpyxl.load_workbook(path)
sheet=workbook.active
rows=sheet.max_row
columns=sheet.max_column
workbook.get_sheet_by_name("Sheeet1")
for r in range(1,rows+1):
    for c in range(1,columns+1):
        values=sheet.cell(rows=r,columns=c)
def utilites():

    driver.implicitly_wait(30)
    driver.maximize_window()

    driver.get('http://atcost.in/')


def searchbar():

    driver.find_element_by_xpath('//input[@placeholder="Search"]').send_keys(values)
    driver.find_element_by_xpath('//i[@class="icon-search button-search hover_icon"]').click()
    driver.find_element_by_xpath('(//button[@class="btn-primary"])[1]').click()
utilites()
searchbar()

